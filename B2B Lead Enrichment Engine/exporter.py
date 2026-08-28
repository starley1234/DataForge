import json
import io
import pandas as pd
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from translit import split_russian_name, get_salutation


def export_to_csv(leads: List[Dict[str, Any]], filepath: str) -> str:
    """Экспорт в CSV в кодировке UTF-8 с BOM (utf-8-sig) для идеального открытия в Microsoft Excel на Windows."""
    df = pd.DataFrame(leads)
    df.to_csv(filepath, index=False, encoding="utf-8-sig")
    return filepath


def export_to_excel(leads: List[Dict[str, Any]], filepath: str) -> str:
    """
    Профессиональный многостраничный экспорт в Excel (.xlsx):
    - Лист 1: «ЛПР и контакты» (подробные данные ЛПР, телефон, email, скоринг)
    - Лист 2: «Сводная аналитика» (KPI базы, топ регионов, статистика)
    - Закрепление строк, автофильтр, стилизация и цветовое кодирование
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ЛПР и контакты"

    if not leads:
        ws.append(["Нет данных"])
        wb.save(filepath)
        return filepath

    column_mappings = [
        ("inn", "ИНН"),
        ("company_name", "Организация"),
        ("dm_full_name", "ФИО ЛПР"),
        ("dm_title", "Должность"),
        ("dm_role_level", "Уровень ЛПР"),
        ("dm_email", "Корпоративный Email"),
        ("email_status", "Статус Email"),
        ("dm_phone", "Телефон"),
        ("dm_phone_type", "Тип телефона"),
        ("phone_carrier", "Оператор связи"),
        ("phone_timezone", "Часовой пояс"),
        ("website", "Сайт"),
        ("region", "Регион"),
        ("okved_name", "Отрасль / ОКВЭД"),
        ("confidence_score", "Скоринг доверия"),
        ("lead_status", "Статус CRM"),
        ("source", "Источник данных"),
        ("notes", "Примечания")
    ]

    # Стилизация
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border_thin = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    fill_even = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    fill_valid = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    font_valid = Font(name="Calibri", size=10, bold=True, color="166534")

    fill_warn = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    font_warn = Font(name="Calibri", size=10, bold=True, color="92400E")

    fill_invalid = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    font_invalid = Font(name="Calibri", size=10, bold=True, color="991B1B")

    # Шапка
    headers = [label for _, label in column_mappings]
    ws.append(headers)
    ws.row_dimensions[1].height = 28

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = cell_border

    # Данные
    keys = [k for k, _ in column_mappings]
    for r_idx, lead in enumerate(leads, start=2):
        row_values = []
        for k in keys:
            val = lead.get(k)
            if k == "confidence_score" and val is not None:
                val = f"{val}%"
            elif val is None or val == "":
                val = "—"
            row_values.append(val)
        ws.append(row_values)
        ws.row_dimensions[r_idx].height = 20

        row_fill = fill_even if r_idx % 2 == 0 else fill_odd
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center")
            cell.fill = row_fill

            if keys[c_idx - 1] == "email_status":
                status = str(cell.value).lower()
                if "valid" in status or "verified" in status:
                    cell.fill = fill_valid
                    cell.font = font_valid
                elif "generated" in status:
                    cell.fill = fill_warn
                    cell.font = font_warn
                elif "invalid" in status or "no_mx" in status or "disposable" in status:
                    cell.fill = fill_invalid
                    cell.font = font_invalid

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(42, max(max_len + 3, 12))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(leads) + 1}"

    # Лист 2: Аналитика
    ws_stats = wb.create_sheet(title="Сводная статистика")
    ws_stats.append(["Метрика базы данных", "Значение"])
    ws_stats.append(["Всего контактов ЛПР", len(leads)])
    
    comp_unique = len(set(l.get("inn") for l in leads if l.get("inn")))
    ws_stats.append(["Уникальных предприятий", comp_unique])
    
    valid_emails = sum(1 for l in leads if l.get("email_status") in ("valid_mx", "verified"))
    ws_stats.append(["Email с действующим MX", valid_emails])
    
    mobiles = sum(1 for l in leads if l.get("dm_phone_type") == "mobile")
    ws_stats.append(["Прямых мобильных телефонов", mobiles])

    ws_stats.row_dimensions[1].height = 24
    for c_idx in (1, 2):
        cell = ws_stats.cell(row=1, column=c_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws_stats.column_dimensions["A"].width = 35
    ws_stats.column_dimensions["B"].width = 20

    wb.save(filepath)
    return filepath


def export_to_json(leads: List[Dict[str, Any]], filepath: str) -> str:
    """Экспорт в форматированный JSON с поддержкой кириллицы."""
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(leads, f, ensure_ascii=False, indent=2)
    return filepath


def export_to_amocrm_csv(leads: List[Dict[str, Any]], filepath: str) -> str:
    """Экспорт в CSV для импорта в amoCRM (сделки и контакты)."""
    amo_rows = []
    for l in leads:
        amo_rows.append({
            "Название сделки": f"{l.get('company_name', 'B2B')} — {l.get('dm_title', 'ЛПР')}",
            "Компания": l.get("company_name", ""),
            "Основной контакт": l.get("dm_full_name", ""),
            "Должность": l.get("dm_title", ""),
            "Рабочий e-mail": l.get("dm_email", ""),
            "Рабочий телефон": l.get("dm_phone", ""),
            "Сайт": l.get("website", ""),
            "Город/Регион": l.get("region", ""),
            "ИНН": l.get("inn", ""),
            "Этап сделки": "Первичный контакт",
            "Примечание": f"Email статус: {l.get('email_status')}, Скоринг: {l.get('confidence_score')}%, Часовой пояс: {l.get('phone_timezone', 'MSK')}"
        })

    df = pd.DataFrame(amo_rows)
    df.to_csv(filepath, index=False, encoding="utf-8-sig")
    return filepath


def export_to_bitrix24_csv(leads: List[Dict[str, Any]], filepath: str) -> str:
    """Экспорт в CSV для импорта лидов в Битрикс24."""
    b24_rows = []
    for l in leads:
        last, first, middle = split_russian_name(l.get("dm_full_name", ""))
        b24_rows.append({
            "Название лида": f"Лид: {l.get('company_name')} ({l.get('dm_full_name')})",
            "Имя": first,
            "Фамилия": last,
            "Отчество": middle,
            "Компания": l.get("company_name", ""),
            "Должность": l.get("dm_title", ""),
            "E-mail": l.get("dm_email", ""),
            "Рабочий телефон": l.get("dm_phone", ""),
            "Сайт": l.get("website", ""),
            "Источник": "B2B Lead Enrichment Engine",
            "Статус": "Не обработан",
            "Комментарий": f"ИНН: {l.get('inn')}, ОКВЭД: {l.get('okved_name')}, Доверие: {l.get('confidence_score')}%, Оператор: {l.get('phone_carrier', '')}"
        })

    df = pd.DataFrame(b24_rows)
    df.to_csv(filepath, index=False, encoding="utf-8-sig")
    return filepath


def export_to_hubspot_csv(leads: List[Dict[str, Any]], filepath: str) -> str:
    """Экспорт в международный формат контактов HubSpot / Salesforce."""
    rows = []
    for l in leads:
        last, first, middle = split_russian_name(l.get("dm_full_name", ""))
        rows.append({
            "First Name": first,
            "Last Name": f"{last} {middle}".strip() if middle else last,
            "Job Title": l.get("dm_title", ""),
            "Email": l.get("dm_email", ""),
            "Phone Number": l.get("dm_phone", ""),
            "Company Name": l.get("company_name", ""),
            "Website URL": l.get("website", ""),
            "City": l.get("region", ""),
            "Lead Status": l.get("lead_status", "NEW"),
            "Tax ID (INN)": l.get("inn", ""),
            "Enrichment Score": l.get("confidence_score", 50)
        })

    df = pd.DataFrame(rows)
    df.to_csv(filepath, index=False, encoding="utf-8")
    return filepath


def export_to_vcard(leads: List[Dict[str, Any]], filepath: str) -> str:
    """
    Экспорт одного или множества контактов в стандартный формат vCard (.vcf 3.0).
    Позволяет импортировать контакты напрямую в iPhone, Android, Apple Contacts, Outlook в один клик.
    """
    vcard_lines = []
    for l in leads:
        last, first, middle = split_russian_name(l.get("dm_full_name", ""))
        vcard_lines.append("BEGIN:VCARD")
        vcard_lines.append("VERSION:3.0")
        vcard_lines.append(f"N:{last};{first};{middle};;")
        vcard_lines.append(f"FN:{l.get('dm_full_name', '')}")
        vcard_lines.append(f"ORG:{l.get('company_name', '')}")
        vcard_lines.append(f"TITLE:{l.get('dm_title', '')}")
        if l.get("dm_email"):
            vcard_lines.append(f"EMAIL;TYPE=INTERNET,WORK:{l['dm_email']}")
        if l.get("dm_phone"):
            p_type = "CELL" if l.get("dm_phone_type") == "mobile" else "WORK"
            vcard_lines.append(f"TEL;TYPE={p_type}:{l['dm_phone']}")
        if l.get("website"):
            vcard_lines.append(f"URL:https://{l['website']}")
        if l.get("region") or l.get("address"):
            vcard_lines.append(f"ADR;TYPE=WORK:;;{l.get('address') or ''};{l.get('region') or ''};;;Russia")
        vcard_lines.append(f"NOTE:ИНН: {l.get('inn', '')} | Скоринг: {l.get('confidence_score', 50)}% | Источник: B2B Lead Engine")
        vcard_lines.append("END:VCARD\n")

    content = "\n".join(vcard_lines)
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(content)
    return filepath


def generate_outreach_email(
    lead: Dict[str, Any],
    offer_type: str = "partnership",
    sender_name: str = "[Ваше Имя]",
    sender_company: str = "[Ваша Компания]",
    sender_title: str = "[Ваша Должность]",
    sender_phone: str = "[Ваш Телефон]"
) -> Dict[str, str]:
    """
    Генерирует готовый персонализированный шаблон холодного B2B-письма:
    - partnership: Стратегическое партнерство и кросс-маркетинг
    - sales: B2B Продажи / Enterprise решение для оптимизации
    - demo: Демонстрация и бесплатный пилотный доступ
    - procurement: Предложение для отдела снабжения и закупок
    - substitution: Импортозамещение и российские аналоги
    - event: Приглашение на закрытый круглый стол / вебинар
    - followup1: Мягкое напоминание (Follow-up 1)
    - followup2: Завершающее письмо (Breakup Email)
    """
    full_name = lead.get("dm_full_name", "")
    comp_name = lead.get("company_name", "Ваша организация")
    title = lead.get("dm_title", "Руководитель")
    email = lead.get("dm_email", "")

    salutation = get_salutation(full_name, formal=True, style="business")

    if offer_type == "sales":
        subject = f"Оптимизация B2B процессов для {comp_name}"
        body = f"""{salutation}

Меня зовут {sender_name}, компания {sender_company}.

Обращаюсь к Вам как к {title.lower()} {comp_name}. Мы разрабатываем и внедряем корпоративные B2B-решения, помогающие предприятиям Вашей сферы сократить издержки и повысить конверсию в продажи.

Изучив масштаб {comp_name}, мы видим высокий потенциал синергии. Недавно мы реализовали аналогичный проект для лидера рынка, увеличив операционную эффективность на 28%.

Будет ли у Вас возможность уделить 15 минут в ближайшие дни (например, во вторник в 11:00 или четверг в 15:00) для краткого ознакомительного онлайн-звонка?

С уважением,
{sender_name}
{sender_title}, {sender_company}
Телефон: {sender_phone}"""

    elif offer_type == "demo":
        subject = f"Демо-доступ и пилотный проект для команды {comp_name}"
        body = f"""{salutation}

Пишу Вам, поскольку Вы курируете ключевые направления развития в {comp_name}.

Мы разработали специализированную систему, которая автоматизирует рутинные процессы и дает руководству полную прозрачность ключевых бизнес-показателей.

Предлагаем организовать персональную 15-минутную демонстрацию возможностей с разбором специфики {comp_name}, а также открыть Вашей команде бесплатный пилотный доступ на 14 дней.

Подскажите, в какой день на этой неделе Вам было бы удобно созвониться?

С уважением,
{sender_name}
{sender_company}
Тел.: {sender_phone}"""

    elif offer_type == "procurement":
        subject = f"Поставки и коммерческое предложение для {comp_name}"
        body = f"""{salutation}

Обращаюсь к Вам по вопросу прямых поставок и участия в закупках компании {comp_name}.

Компания {sender_company} является производителем и прямым поставщиком продукции. Мы предлагаем гибкие условия оплаты, соблюдение сроков по SLA и персональную систему скидок до 18% для постоянных партнеров.

Хотели бы направить наш актуальный каталог с оптовым прайс-листом и рассчитать тестовую спецификацию.

Подскажите, пожалуйста, с кем из специалистов отдела снабжения можно связаться для передачи коммерческого предложения?

С уважением,
{sender_name}
{sender_company}
Контакты: {sender_phone}"""

    elif offer_type == "substitution":
        subject = f"Российская альтернатива и импортозамещение для {comp_name}"
        body = f"""{salutation}

Пишу Вам по вопросу перехода на доверенные отечественные решения в инфраструктуре {comp_name}.

{sender_company} разработала полностью независимый программно-аппаратный комплекс, включенный в реестр Минцифры РФ. Наше решение успешно замещает зарубежные аналоги (ZoomInfo, SAP, Salesforce) без потери функционала и с экономией до 40% бюджета.

Будем рады предоставить кейсы миграции и рассчитать стоимость внедрения для {comp_name}.

Удобно ли запланировать 15-минутный звонок для обсуждения технических деталей?

С уважением,
{sender_name}
{sender_company}
Телефон: {sender_phone}"""

    elif offer_type == "event":
        subject = f"Приглашение: закрытый экспертный круглый стол для руководства {comp_name}"
        body = f"""{salutation}

Приглашаем Вас и ведущих экспертов {comp_name} принять участие в закрытом онлайн-митапе «Тренды цифровизации и эффективность B2B-бизнеса в 2026 году».

В программе:
• Разбор кейсов лидеров отрасли
• Автоматизация и искусственный интеллект в продажах
• Сессия вопросов и нетворкинг с топ-менеджерами

Участие для представителей {comp_name} бесплатное по предварительной регистрации. 

Будем рады выслать персональное приглашение и ссылку на трансляцию.

С уважением,
{sender_name}
{sender_company}"""

    elif offer_type == "followup1":
        subject = f"Re: Сотрудничество с {comp_name}"
        body = f"""{salutation}

Писал Вам на прошлой неделе по поводу совместного проекта между {sender_company} и {comp_name}.

Понимаю Вашу высокую занятость, поэтому просто хотел кратко напомнить о себе. 

Мы уверены, что наше решение принесет ощутимую пользу Вашему бизнесу. Если вопрос актуален, с удовольствием подстроюсь под Ваш график для 10-минутного созвона.

Хорошего рабочего дня!

С уважением,
{sender_name}
{sender_company}
{sender_phone}"""

    elif offer_type == "followup2":
        subject = f"Финальное уточнение // {comp_name}"
        body = f"""{salutation}

Так как не получил ответа на предыдущие письма, полагаю, что сейчас вопрос оптимизации процессов не в фокусе приоритетов {comp_name}.

Больше не стану беспокоить Вас письмами. Если ситуация изменится и Вам потребуется надежное B2B-решение, Вы всегда можете связаться со мной по этому адресу или телефону {sender_phone}.

Желаю успехов и процветания Вашему бизнесу!

С уважением,
{sender_name}
{sender_company}"""

    else:  # partnership
        subject = f"Предложение о стратегическом партнерстве с {comp_name}"
        body = f"""{salutation}

Обращаюсь к Вам по вопросу стратегического сотрудничества с {comp_name}.

Компания {sender_company} реализует комплексные B2B-проекты. Мы видим отличные точки соприкосновения и взаимную синергию между нашими продуктами и масштабом Вашего бизнеса.

Хотели бы направить краткую презентацию концепции партнерства и обсудить взаимовыгодные форматы взаимодействия.

Буду признателен за обратную связь или подсказку, с кем из Ваших коллег было бы целесообразно созвониться по данному вопросу.

С уважением,
{sender_name}
{sender_title}, {sender_company}
Телефон: {sender_phone}"""

    return {
        "recipient_email": email,
        "recipient_name": full_name,
        "subject": subject,
        "body": body
    }


def generate_cold_calling_script(lead: Dict[str, Any]) -> Dict[str, Any]:
    """
    Генерирует готовый сценарий холодного телефонного звонка (Cold Calling Script):
    - Секретарский барьер (Gatekeeper Bypass)
    - Открытие разговора с ЛПР (30-sec Pitch Hook)
    - Отработка 5 типовых B2B возражений
    - Закрытие на онлайн-встречу
    """
    full_name = lead.get("dm_full_name", "")
    last, first, middle = split_russian_name(full_name)
    comp_name = lead.get("company_name", "компании")
    title = lead.get("dm_title", "руководитель")

    address_fio = f"{first} {middle}".strip() if (first and middle) else (first or full_name)

    return {
        "lead_name": full_name,
        "company_name": comp_name,
        "title": title,
        "gatekeeper_script": f"«Здравствуйте! Соедините, пожалуйста, с {address_fio} по вопросу оптимизации поставок для {comp_name}. Он(а) ожидает звонка по совместному проекту.»",
        "intro_pitch": f"«{address_fio}, добрый день! Меня зовут [Имя], компания [Компания]. Звоню кратко по делу: мы разработали B2B-решение для компаний Вашей отрасли, которое сокращает расходы на закупки на 15-20%. Уделите 1 минуту, чтобы понять, насколько это применимо в {comp_name}? »",
        "objections": [
            {
                "objection": "«Нам ничего не нужно / Все устраивает»",
                "answer": "«Отлично понимаю Вас, многие наши текущие партнеры сначала говорили так же, пока не сравнили цифры эффективности. Мы не предлагаем прямо сейчас что-то менять — давайте просто покажем кейс аналогичной компании за 10 минут в Zoom, а выводы сделаете сами.»"
            },
            {
                "objection": "«Отправьте предложение на общую почту (info@)»",
                "answer": "«Обязательно отправлю, но на info@ письма часто теряются среди сотен сообщений. Чтобы не тратить Ваше время на шаблонный файл из 30 страниц, могу я уточнить 1 ключевой параметр и отправить расчет лично Вам в WhatsApp или Telegram?»"
            },
            {
                "objection": "«У нас уже есть надежный поставщик / подрядчик»",
                "answer": "«Это замечательно, мы и не предлагаем сразу разрывать отношения с текущим партнером. Мы предлагаем иметь запасной вариант и альтернативный расчет. В случае непредвиденных задержек или изменения цен у Вас всегда будет готовая опция.»"
            },
            {
                "objection": "«Сейчас нет бюджета / кризис»",
                "answer": "«Именно поэтому мы предлагаем решение, которое окупается за первый же месяц и помогает высвободить замороженные средства. Давайте проведем аудит без каких-либо финансовых обязательств с Вашей стороны?»"
            },
            {
                "objection": "«Я этим не занимаюсь / перезвоните позже»",
                "answer": f"«Понял Вас, {address_fio}. Подскажите, пожалуйста, кто в {comp_name} отвечает за это направление, чтобы я связался напрямую и не отвлекал Вас?»"
            }
        ],
        "closing": "«Давайте поступим так: во вторник в 11:00 или в четверг в 15:00 созвонимся на 15 минут в онлайн-формате, покажем готовый расчет для Вашей компании. Какое время Вам удобнее?»"
    }
