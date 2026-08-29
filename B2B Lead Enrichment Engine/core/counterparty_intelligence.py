"""
Модуль аналитики и проверки контрагентов Rusprofile 360° (Counterparty Intelligence).
Обеспечивает формирование полного досье благонадежности, расчет матрицы рисков,
проверку по 38 государственным реестрам РФ и выгрузку отчетов о должной осмотрительности.
"""

import io
from typing import Dict, Any, Optional
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from sources.rusprofile_sources import CounterpartyDataAggregator
from core.engine import EnrichmentEngine


class CounterpartyIntelligenceEngine:
    def __init__(self, engine: Optional[EnrichmentEngine] = None):
        self.engine = engine or EnrichmentEngine()
        self.aggregator = CounterpartyDataAggregator()

    def get_full_dossier(self, query: str) -> Optional[Dict[str, Any]]:
        """
        Возвращает полное досье организации по ИНН, ОГРН или наименованию.
        Сначала ищет базовые данные в EnrichmentEngine / ЕГРЮЛ, затем строит аналитическое досье.
        """
        clean_q = str(query).strip()
        if not clean_q:
            return None

        base_comp = self.engine.fetch_and_enrich(clean_q, scrape_web=False, verify_emails=False)
        base_dict = None
        if base_comp:
            base_dict = {
                "name": base_comp.name,
                "short_name": base_comp.short_name,
                "inn": base_comp.inn,
                "ogrn": base_comp.ogrn,
                "kpp": base_comp.kpp,
                "region": base_comp.region,
                "city": base_comp.city,
                "address": base_comp.address,
                "okved": base_comp.okved,
                "okved_name": base_comp.okved_name,
                "revenue_rub": base_comp.revenue_rub,
                "employees_count": base_comp.employees_count,
                "website": base_comp.website,
                "domain": base_comp.domain,
                "decision_makers": base_comp.decision_makers
            }

        inn_to_use = base_comp.inn if base_comp else clean_q
        dossier = self.aggregator.build_full_dossier(inn_to_use, base_dict)

        # Добавляем найденных ЛПР и контакты из нашей БД
        leads = self.engine.get_all_leads(query=inn_to_use)
        dossier["leads"] = leads

        return dossier

    def generate_due_diligence_report_md(self, dossier: Dict[str, Any]) -> str:
        """
        Генерирует официальное заключение о проверке контрагента (Отчет о должной осмотрительности) в Markdown.
        """
        s = dossier["summary"]
        l = dossier["leadership"]
        f = dossier["finance"]
        p = dossier["procurement"]
        c = dossier["courts"]
        rf = dossier["risk_factors"]

        md = f"""# Отчет о проверке контрагента и должной осмотрительности (Due Diligence 360°)
**Дата формирования:** {dossier.get('generated_at', '2026-08-29')}
**Стандарт проверки:** Ст. 54.1 НК РФ, Регламенты ФНС РФ, методология Rusprofile

---

## 1. Общие сведения и реквизиты организации
- **Полное наименование:** {s['name']}
- **Краткое наименование:** {s['short_name']}
- **ИНН / КПП:** `{s['inn']}` / `{s['kpp']}`
- **ОГРН:** `{s['ogrn']}` (Дата регистрации: {s['registration_date']}, возраст бизнеса: {s['age_years']} лет)
- **Юридический адрес:** {s['address']} (Массовый адрес: {'ДА ⚠️' if s['is_mass_address'] else 'НЕТ (надежно)'})
- **Статус ЕГРЮЛ:** **{s['status_text']}**
- **Уставный капитал:** {s['capital_rub']:,.0f} руб.
- **Основной вид деятельности (ОКВЭД):** {s['okved']} — {s['okved_name']}
- **Налоговый орган:** {s['tax_authority']}
- **Система налогообложения:** {s['tax_system']}
- **Среднесписочная численность:** {s['employees_count']} сотрудников

---

## 2. Руководство и Учредители
- **Руководитель:** {l['ceo_title']} **{l['ceo_name']}** (ИНН: `{l['ceo_inn']}`)
  - В реестре дисквалифицированных лиц: {'⚠️ ДА (СТОП-ФАКТОР)' if l['is_disqualified'] else 'НЕТ'}
  - Массовый руководитель: {'⚠️ ДА (>5 юрлиц)' if l['is_mass_director'] else 'НЕТ'}
- **Учредители:**
"""
        for founder in dossier["founders"]:
            md += f"  - **{founder['name']}** — доля {founder['share_percent']}% ({founder['share_rub']:,.0f} руб.)\n"

        md += f"""
---

## 3. Финансовое состояние (ГИР БО ФНС РФ)
- **Выручка за последний отчетный год:** {f['revenue_latest']:,.0f} руб.
- **Чистая прибыль:** {f['profit_latest']:,.0f} руб.
- **Активы баланса:** {f['assets_latest']:,.0f} руб.
- **Чистые активы:** {f['net_assets']:,.0f} руб.
- **Уплачено налогов и сборов:** {f['taxes_paid_total']:,.0f} руб. (НДС: {f['taxes_breakdown']['vat']:,.0f} руб., Налог на прибыль: {f['taxes_breakdown']['income_tax']:,.0f} руб.)
- **Задолженность по налогам:** {f['tax_debt']:,.0f} руб. {'(Имеется задолженность ⚠️)' if f['has_tax_debt'] else '(Отсутствует)'}
- **Приостановления по счетам (блокировки ФНС):** {f['account_blocks_count']} решений

---

## 4. Госзакупки и надежность поставщика (44-ФЗ / 223-ФЗ / РНП)
- **Заключено госконтрактов в роли Поставщика:** {p['supplier_contracts_count']} контрактов на сумму **{p['supplier_contracts_sum']:,.0f} руб.**
- **Реестр недобросовестных поставщиков (РНП ФАС):** **{p['rnp_status']}**

---

## 5. Судебная практика и исполнительные производства
- **Арбитражные дела в роли Истца:** {c['plaintiff_count']} дел (на сумму {c['plaintiff_sum']:,.0f} руб.)
- **Арбитражные дела в роли Ответчика:** {c['defendant_count']} дел (на сумму {c['defendant_sum']:,.0f} руб.)
- **Исполнительные производства ФССП:** {dossier['fssp']['active_proceedings_count']} активных производств на сумму {dossier['fssp']['active_debt_sum']:,.0f} руб.

---

## 6. Итоговая оценка надежности и матрица рисков
### **Индекс надежности Rusprofile:** `{s['reliability_score']} / 100` ({s['reliability_text']})

### 🟢 Положительные факторы:
"""
        for pos in rf["positive"]:
            md += f"- ✓ {pos}\n"

        if rf["warnings"]:
            md += "\n### 🟡 Факторы, требующие внимания:\n"
            for w in rf["warnings"]:
                md += f"- ⚠️ {w}\n"

        if rf["critical"]:
            md += "\n### 🔴 Критические факторы риска (СТОП-ФАКТОРЫ):\n"
            for crit in rf["critical"]:
                md += f"- ❌ {crit}\n"

        md += "\n---\n*Заключение сформировано автоматически системой DataForge Lead Intelligence Enterprise по открытым источникам ФНС, ФССП, ЕИС и КАД.*"
        return md

    def export_due_diligence_excel(self, dossier: Dict[str, Any], output_path: str) -> str:
        """
        Экспортирует стилизованный официальный отчет о проверке контрагента в Excel.
        """
        s = dossier["summary"]
        l = dossier["leadership"]
        f = dossier["finance"]
        p = dossier["procurement"]
        c = dossier["courts"]
        rf = dossier["risk_factors"]

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # 1. Лист Досье
            dossier_data = [
                ("ИНН", s["inn"]),
                ("КПП", s["kpp"]),
                ("ОГРН", s["ogrn"]),
                ("Полное наименование", s["name"]),
                ("Краткое наименование", s["short_name"]),
                ("Статус", s["status_text"]),
                ("Дата регистрации", s["registration_date"]),
                ("Возраст бизнеса (лет)", s["age_years"]),
                ("Уставный капитал (руб)", s["capital_rub"]),
                ("Юридический адрес", s["address"]),
                ("Массовый адрес регистрации", "Да" if s["is_mass_address"] else "Нет"),
                ("Руководитель", f"{l['ceo_title']} {l['ceo_name']}"),
                ("Дисквалификация руководителя", "Да" if l["is_disqualified"] else "Нет"),
                ("Основной ОКВЭД", f"{s['okved']} {s['okved_name']}"),
                ("Выручка 2025 (руб)", f["revenue_latest"]),
                ("Чистая прибыль 2025 (руб)", f["profit_latest"]),
                ("Чистые активы (руб)", f["net_assets"]),
                ("Госконтрактов (шт)", p["supplier_contracts_count"]),
                ("Сумма госконтрактов (руб)", p["supplier_contracts_sum"]),
                ("Реестр недобросовестных поставщиков (РНП)", p["rnp_status"]),
                ("Арбитражные дела в роли ответчика (руб)", c["defendant_sum"]),
                ("Долги ФССП (руб)", dossier["fssp"]["active_debt_sum"]),
                ("Индекс надежности (0-100)", s["reliability_score"]),
                ("Категория надежности", s["reliability_text"])
            ]
            df_dossier = pd.DataFrame(dossier_data, columns=["Параметр проверки", "Значение"])
            df_dossier.to_excel(writer, sheet_name="Досье контрагента", index=False)

            # 2. Лист Финансовая динамика
            df_fin = pd.DataFrame(f["history"])
            df_fin.columns = ["Год", "Выручка (руб)", "Чистая прибыль (руб)", "Активы (руб)"]
            df_fin.to_excel(writer, sheet_name="Финансы (ГИР БО)", index=False)

            # 3. Лист Факторы рисков
            risk_rows = []
            for item in rf["positive"]:
                risk_rows.append({"Тип фактора": "🟢 Положительный факт", "Описание": item})
            for item in rf["warnings"]:
                risk_rows.append({"Тип фактора": "🟡 Требует внимания", "Описание": item})
            for item in rf["critical"]:
                risk_rows.append({"Тип фактора": "🔴 Стоп-фактор / Риск", "Описание": item})
            df_risks = pd.DataFrame(risk_rows)
            df_risks.to_excel(writer, sheet_name="Матрица рисков ФНС", index=False)

            # Стилизация книги Excel
            wb = writer.book
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

            for ws in wb.worksheets:
                for col in ws.iter_cols(min_row=1, max_row=1):
                    for cell in col:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(18, min(max_len + 4, 60))
        return output_path
